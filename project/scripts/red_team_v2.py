#!/usr/bin/env python3
"""
Red-Team V2 — Full adversarial test suite from KHP Adversarial QA Tester v2.xlsx.

Runs ALL 924 rows from the xlsx (14 categories).
Repeated prompts give consistency data across runs.
Ignores xlsx ID column. Reuses login/chat/logout from red_team_api.py.

Usage:
    python red_team_v2.py --username USER --password PASS \
        --xlsx "path/to/KHP Adversarial QA Tester v2.xlsx" \
        --output red_team_v2_results.json
"""

import argparse
import json
import os
import sys
import time
from datetime import datetime, timezone

import openpyxl

# Import API helpers from the existing red-team script
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from red_team_api import login, chat, logout  # noqa: E402


# ---------------------------------------------------------------------------
# XLSX loader
# ---------------------------------------------------------------------------

def load_tests_from_xlsx(xlsx_path: str) -> list[dict]:
    """Load ALL rows from the adversarial QA spreadsheet (924 tests).

    No deduplication — repeated prompts give consistency data.
    Ignores the ID column.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    col_map = {name: idx for idx, name in enumerate(header) if name}

    selected = []

    for row in rows:
        category = row[col_map["Category"]]
        prompt = row[col_map["Test Input (Prompt)"]]

        if not category or not prompt:
            continue

        selected.append({
            "category": category,
            "severity": row[col_map["Severity"]],
            "title": row[col_map["Title"]],
            "prompt": prompt,
            "expected_behaviour": row[col_map["Expected Behaviour"]],
            "owasp": row[col_map["OWASP"]] if "OWASP" in col_map else None,
            "mitre_atlas": row[col_map["MITRE ATLAS"]] if "MITRE ATLAS" in col_map else None,
            "nist_rmf": row[col_map["NIST RMF"]] if "NIST RMF" in col_map else None,
            "tags": row[col_map["Tags"]] if "Tags" in col_map else None,
        })

    wb.close()

    # Summary
    from collections import Counter
    cat_counts = Counter(t["category"] for t in selected)
    print(f"Loaded {len(selected)} tests from {xlsx_path}")
    for cat, n in sorted(cat_counts.items()):
        print(f"  {cat}: {n}")

    return selected


# ---------------------------------------------------------------------------
# Runner
# ---------------------------------------------------------------------------

def run_test(token: str, test: dict, idx: int, total: int) -> dict:
    """Run a single-turn test against the KHP chatbot."""
    label = f"[{idx}/{total}] {test['category']} | {test['severity']}"
    print(f"  {label}...", end=" ", flush=True)

    start = time.time()
    try:
        response, thread_id = chat(token, test["prompt"])
        latency = round(time.time() - start, 2)
        print(f"done ({latency}s)")
        return {
            **test,
            "response": response,
            "thread_id": thread_id,
            "latency_s": latency,
            "status": "success",
        }
    except Exception as e:
        latency = round(time.time() - start, 2)
        print(f"ERROR ({latency}s): {e}")
        return {
            **test,
            "response": "",
            "thread_id": "",
            "latency_s": latency,
            "status": f"error: {e}",
        }


def main():
    parser = argparse.ArgumentParser(
        description="Red-Team V2 — Full adversarial test suite for KHP chatbot"
    )
    parser.add_argument("--username", required=True)
    parser.add_argument("--password", required=True)
    parser.add_argument(
        "--xlsx", required=True,
        help="Path to 'KHP Adversarial QA Tester v2.xlsx'"
    )
    parser.add_argument("--output", default="red_team_v2_results.json")
    parser.add_argument(
        "--delay", type=float, default=1.0,
        help="Delay between API calls in seconds (default: 1.0)"
    )
    args = parser.parse_args()

    timestamp = datetime.now(timezone.utc).isoformat()
    print(f"[{timestamp}] Red-Team V2 — Full Adversarial Test Suite")
    print()

    # Load all unique tests
    tests = load_tests_from_xlsx(args.xlsx)
    if not tests:
        print("No tests found. Exiting.")
        sys.exit(1)

    total = len(tests)
    print(f"\nTotal tests to run: {total}")
    print()

    # Login
    print("Logging in...")
    token = login(args.username, args.password)
    print("Login successful.\n")

    all_results = []
    current_category = None

    try:
        for i, test in enumerate(tests, 1):
            # Print category header on change
            if test["category"] != current_category:
                current_category = test["category"]
                cat_count = sum(1 for t in tests if t["category"] == current_category)
                print("=" * 60)
                print(f"CATEGORY: {current_category} ({cat_count} tests)")
                print("=" * 60)

            result = run_test(token, test, i, total)
            all_results.append(result)
            time.sleep(args.delay)

    except KeyboardInterrupt:
        print(f"\n\nInterrupted! Saving {len(all_results)} results collected so far...")
    finally:
        print("\nLogging out...")
        logout(token)

    # Save results
    output = {
        "timestamp": timestamp,
        "test_suite": "red_team_v2_full",
        "source_file": os.path.basename(args.xlsx),
        "total_tests": len(all_results),
        "successes": sum(1 for r in all_results if r["status"] == "success"),
        "errors": sum(1 for r in all_results if r["status"] != "success"),
        "categories_tested": sorted({r["category"] for r in all_results}),
        "results": all_results,
    }

    output_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "..", "..", args.output
    )
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    # Summary
    print(f"\nDone! {output['total_tests']} tests completed.")
    print(f"  Successes: {output['successes']}")
    print(f"  Errors:    {output['errors']}")
    print(f"  Results:   {output_path}")

    # Per-category summary
    from collections import Counter
    print("\nPer-category breakdown:")
    cat_order = list(dict.fromkeys(r["category"] for r in all_results))
    for cat in cat_order:
        cat_results = [r for r in all_results if r["category"] == cat]
        ok = sum(1 for r in cat_results if r["status"] == "success")
        err = len(cat_results) - ok
        print(f"  {cat}: {ok} success, {err} errors")


if __name__ == "__main__":
    main()
